import logging

import xlwt
from openpyxl import load_workbook
import os

def append_to_xlsx(filename, svod_rows, count_rows, error_rows):
    wb = load_workbook(filename)
    ws = wb['Исходящая ПСО сводный']

def exportDataxls(rows, date, newdate, rows1, svod_rows, error_rows):


    wb = xlwt.Workbook(encoding='utf-8')
    #Сводный отчет
    ws_svod = wb.add_sheet('Исходящая ПСО сводный')
    new_row_num = 0
    borders_svod = xlwt.Borders()
    borders_svod.bottom = 1
    borders_svod.left = 1
    borders_svod.right = 1
    borders_svod.top = 1
    borders_svod.bottom_colour = 0
    borders_svod.right_colour = 0
    borders_svod.left_colour = 0
    borders_svod.top_colour = 0
    font_style_svod = xlwt.XFStyle()
    font_svod = xlwt.Font()
    font_svod.bold = False
    font_style_svod.font = font_svod
    style_string_svod = "font: bold on; borders: bottom 1; borders: left 1; borders: right 1; borders: top 1"
    style = xlwt.easyxf(style_string_svod)
    columns_svod = ["Номер документа", "Наименование журнала регистрации", "Отправитель", "Департамент отправителя", "Дата и время отправки",
                    "Кол-во отправленных пакетов",
               "Зарегистрирован", "Отправлен","Доставлен","Ошибка на стороне получателя","Ошибка отправки","Ошибка выгрузки"]

    font_style_svod.borders = borders_svod
    for col_num in range(len(columns_svod)):
        ws_svod.write(new_row_num, col_num, columns_svod[col_num], style)
        cells = xlwt.Column(col_num, ws_svod)
        cells.width = 5000
    for row in svod_rows:
        new_row_num += 1
        a = [row[0],row[1],row[2],row[3],str(row[4]),row[5],row[6], row[7],row[8],row[9],row[10],row[11]]

        for col_num in range(len(a)):
            ws_svod.col(0).width = 30 * 256
            ws_svod.col(1).width = 50 * 256
            ws_svod.col(2).width = 40 * 256
            ws_svod.col(3).width = 40 * 256
            ws_svod.col(4).width = 30 * 256
            ws_svod.col(5).width = 10 * 256
            ws_svod.col(6).width = 10 * 256
            ws_svod.col(7).width = 10 * 256
            ws_svod.col(8).width = 10 * 256
            ws_svod.col(9).width = 10 * 256
            ws_svod.col(10).width = 10 * 256
            ws_svod.col(11).width = 10 * 256
            if int(row[5]) != int(row[6]):
                ws_svod.write(new_row_num, col_num, a[col_num], style)
            else:ws_svod.write(new_row_num, col_num, a[col_num], font_style_svod)
    print(f'запись сводного отчета кол-во строк: {len(svod_rows)}' )
    logging.info(f'запись сводного отчета кол-во строк: {len(svod_rows)}')
    # подробный отчет
    ws = wb.add_sheet('Исходящая ПСО подробный')

    # Sheet header, first row
    new_row_num = 0
    borders = xlwt.Borders()
    borders.bottom = 1
    borders.left = 1
    borders.right = 1
    borders.top = 1
    borders.bottom_colour = 0
    borders.right_colour = 0
    borders.left_colour = 0
    borders.top_colour = 0
    font_style = xlwt.XFStyle()
    font = xlwt.Font()
    font.bold = True
    font_style.font = font

    style_string = "font: bold on; borders: bottom 1; borders: left 1; borders: right 1; borders: top 1"
    style = xlwt.easyxf(style_string)

    columns = ['Статус выгрузки','Отправитель', 'департамент отправителя','Получатель', 'Департамент получателя','Номер документа','Наименование журнала регистрации',
               'Текущий статус пакета','Дата и время постановки в очередь']
    font_style = xlwt.XFStyle()

    font_style.borders = borders
    columns_count = ['Количество пакетов', 'статус']
    for col_num1 in range(len(columns_count)):
        ws.write(new_row_num, col_num1, columns_count[col_num1], style)
        cells = xlwt.Column(col_num1, ws)
        cells.width = 500
    summa = 0
    for row1 in rows1:
        ro1 = list(row1)
        new_row_num += 1

        if ro1[1]== None:

            ro1[1] = "Ошибка выгрузки"

        b = [row1[0],ro1[1]]

        summa = summa +row1[0]
        ws.write(new_row_num, 0, b[0], style)
        ws.write(new_row_num, 1, b[1], font_style)
    ws.write(new_row_num+1, 0, summa, style)
    ws.write(new_row_num+1, 1, 'Всего пакетов', font_style)
    row_num = new_row_num + 3
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], style)

    # Sheet body, remaining rows
        cells = xlwt.Column(col_num,ws)
        cells.width = 5000

    for row in rows:
        row_num += 1
        a = [row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7], str(row[8]) ]

        for col_num in range(len(a)):
            ws.col(0).width = 30 * 256
            ws.col(1).width = 50 * 256
            ws.col(2).width = 30 * 256
            ws.col(3).width = 90 * 256
            ws.col(4).width = 60 * 256
            ws.col(5).width = 30 * 256
            ws.col(6).width = 30 * 256
            ws.col(7).width = 30 * 256
            ws.col(8).width = 30 * 256
            ws.write(row_num, col_num, a[col_num], font_style)
        #
        #ws.write(row_num, , [[row.srvaddr],[row.namedisk], [str(row.freespace)],[str(row.allspace)],[row.date_up]], font_style)
    print(f'Запись подробного отчета, кол-во строк : {len(rows)}' )
    logging.info(f'Запись подробного отчета, кол-во строк : {len(rows)}')
    #error book
    ws_error = wb.add_sheet('Исходящая ПСО ошибки')

    # Sheet header, first row
    new_row_num_error = 0
    borders_error = xlwt.Borders()
    borders_error.bottom = 1
    borders_error.left = 1
    borders_error.right = 1
    borders_error.top = 1
    borders_error.bottom_colour = 0
    borders_error.right_colour = 0
    borders_error.left_colour = 0
    borders_error.top_colour = 0
    font_style_error = xlwt.XFStyle()
    font_error = xlwt.Font()
    font_error.bold = True
    font_style_error.font = font_error

    style_string_error = "font: bold on; borders: bottom 1; borders: left 1; borders: right 1; borders: top 1"
    style_error = xlwt.easyxf(style_string_error)

    columns_error = ['Статус выгрузки', 'Отправитель',"Департамент отправителя", 'Получатель','Департамент получателя', 'Номер документа', 'Наименование журнала регистрации',
               'Текущий статус пакета', 'Дата и время постановки в очередь']
    font_style_error = xlwt.XFStyle()

    font_style_error.borders = borders

    for col_num_error in range(len(columns)):
        ws_error.write(new_row_num_error, col_num_error, columns_error[col_num_error], style_error)

        # Sheet body, remaining rows
        cells_error = xlwt.Column(col_num_error, ws_error)

        cells_error.width = 5000

    for row_error in error_rows:
        new_row_num_error += 1
        a_error = [row_error[0], row_error[1], row_error[2], row_error[3], row_error[4], row_error[5], row_error[6], row_error[7],str(row_error[8])]

        for col_num_error in range(len(a)):
            ws_error.col(0).width = 30 * 256
            ws_error.col(1).width = 50 * 256
            ws_error.col(2).width = 30 * 256
            ws_error.col(3).width = 90 * 256
            ws_error.col(4).width = 60 * 256
            ws_error.col(5).width = 30 * 256
            ws_error.col(6).width = 30 * 256
            ws_error.col(7).width = 30 * 256
            ws_error.col(8).width = 30 * 256
            ws_error.write(new_row_num_error, col_num_error, a_error[col_num_error], font_style_error)
        #
        # ws.write(row_num, , [[row.srvaddr],[row.namedisk], [str(row.freespace)],[str(row.allspace)],[row.date_up]], font_style)
    print(f'Запись отчета по ошибочным пакетам, кол-во строк: {len(error_rows)}' )
    logging.info(f'Запись отчета по ошибочным пакетам, кол-во строк: {len(error_rows)}')
    try:
        os.mkdir(".\\reports\\")

    except: print('папка с отчетами уже существует, не создаём')
    wb.save(
        f".\\reports\\Отчет ПСО отправленные док-ты-(с {newdate.strftime('%Y-%m-%d')} по {date.strftime('%Y-%m-%d')}).xls")
def create_report_APE(rows, date):


    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Документы без резолюции')

    # Sheet header, first row
    new_row_num = 0
    borders = xlwt.Borders()
    borders.bottom = 1
    borders.left = 1
    borders.right = 1
    borders.top = 1
    borders.bottom_colour = 0
    borders.right_colour = 0
    borders.left_colour = 0
    borders.top_colour = 0
    font_style = xlwt.XFStyle()
    font = xlwt.Font()
    font.bold = True
    font_style.font = font

    style_string = "font: bold on; borders: bottom 1; borders: left 1; borders: right 1; borders: top 1"
    style = xlwt.easyxf(style_string)

    columns = ['Наименование журнала регистрации','рег. дата документа','рег. номер документа']
    font_style = xlwt.XFStyle()

    font_style.borders = borders

    row_num = new_row_num
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], style)

    # Sheet body, remaining rows
        cells = xlwt.Column(col_num,ws)
        cells.width = 5000

    for row in rows:
        row_num += 1
        a = [row[0],str(row[1]),row[2]]

        for col_num in range(len(a)):
            ws.col(0).width = 30 * 256
            ws.col(1).width = 100 * 256
            ws.col(2).width = 30 * 256
            ws.write(row_num, col_num, a[col_num], font_style)
        #
        #ws.write(row_num, , [[row.srvaddr],[row.namedisk], [str(row.freespace)],[str(row.allspace)],[row.date_up]], font_style)
    print('start adding new table', row_num)


    wb.save(f".\\reports\\Список документов без резолюции Артюхин Р.Е. {date.strftime('%Y-%m-%d')}.xls")